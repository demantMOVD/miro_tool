"""
Timeline Excel Generator
Generates a team timeline spreadsheet with:
  - Sheet 1: Info/Config page (variables)
  - Sheet 2: Timeline view (iterations over X months, team row + member rows)
"""

import openpyxl
import os as _os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONSTANTS / DEFAULTS
# ─────────────────────────────────────────────
_SCRIPT_DIR = _os.path.dirname(_os.path.abspath(__file__))
OUTPUT_FILE = _os.path.join(_SCRIPT_DIR, "TeamTimeline.xlsx")

DEFAULT_PROJECT_NAME   = "My Project"
DEFAULT_START_MONTH    = "2026-04"      # YYYY-MM
DEFAULT_NUM_MONTHS     = 6
DEFAULT_TEAM_NAME      = "Team Alpha"
DEFAULT_MEMBERS        = ["Alice", "Bob", "Charlie", "Diana"]

WEEKS_PER_ITER         = 2             # constant – each block = 2 weeks (1 iteration)

# Palette
COL_HEADER_FILL   = "1F3864"   # dark navy
COL_HEADER_FONT   = "FFFFFF"
COL_TEAM_FILL     = "2E75B6"   # mid-blue  (team row)
COL_TEAM_FONT     = "FFFFFF"
COL_MEMBER_FILL   = "BDD7EE"   # light-blue (member rows)
COL_MEMBER_FONT   = "1F3864"
COL_ITER_EVEN     = "DDEBF7"   # iteration column even shading
COL_ITER_ODD      = "EBF3FB"
COL_INFO_LABEL    = "1F3864"
COL_INFO_VALUE    = "BDD7EE"
COL_ACCENT        = "2E75B6"
COL_TASK_FILL     = "FFF2CC"   # task cell background
COL_TASK_BORDER   = "BF9000"


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def make_border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def styled_cell(ws, row, col, value=None, bold=False, font_color="000000",
                fill_color=None, align_h="center", align_v="center",
                wrap=False, font_size=11, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v,
                               wrap_text=wrap)
    if fill_color:
        cell.fill = make_fill(fill_color)
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def calculate_iterations(start_month_str, num_months):
    """Return list of (label, start_date_str) for each 2-week iteration."""
    from datetime import date, timedelta
    year, month = map(int, start_month_str.split("-"))
    start = date(year, month, 1)

    # total days ≈ num_months * 30.44
    import math
    total_days = math.ceil(num_months * 30.4375)
    end = start + timedelta(days=total_days)

    iters = []
    current = start
    iter_num = 1
    while current < end:
        iter_end = current + timedelta(days=13)
        label = f"Iter {iter_num}\n{current.strftime('%d %b')}–{min(iter_end,end).strftime('%d %b')}"
        iters.append((label, current.strftime("%Y-%m-%d"), iter_num))
        current += timedelta(weeks=WEEKS_PER_ITER)
        iter_num += 1
    return iters


# ─────────────────────────────────────────────
#  SHEET 1 – INFO / CONFIG
# ─────────────────────────────────────────────
def build_info_sheet(wb, members, project_name, start_month, num_months, team_name):
    ws = wb.active
    ws.title = "Info & Config"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 3

    b = make_border("thin", "B8CCE4")

    # ── Title ──────────────────────────────
    ws.merge_cells("B2:F2")
    styled_cell(ws, 2, 2, "📋  TIMELINE CONFIGURATOR", bold=True,
                font_color=COL_HEADER_FONT, fill_color=COL_HEADER_FILL,
                font_size=16, border=b)
    ws.row_dimensions[2].height = 36

    # ── Project settings ───────────────────
    def section_title(row, text):
        ws.merge_cells(f"B{row}:F{row}")
        styled_cell(ws, row, 2, text, bold=True,
                    font_color=COL_HEADER_FONT, fill_color=COL_ACCENT,
                    font_size=12, border=b)
        ws.row_dimensions[row].height = 22

    def config_row(row, label, value, note=""):
        styled_cell(ws, row, 2, label, bold=True,
                    font_color=COL_HEADER_FONT, fill_color=COL_INFO_LABEL,
                    align_h="left", border=b)
        cell = styled_cell(ws, row, 3, value,
                           fill_color=COL_INFO_VALUE, align_h="left", border=b)
        cell.font = Font(color="1F3864", size=11)
        if note:
            styled_cell(ws, row, 5, note, align_h="left",
                        font_color="7F7F7F", font_size=9)
        ws.row_dimensions[row].height = 20
        return cell

    section_title(4, "  PROJECT SETTINGS")
    config_row(5, "Project Name",   project_name,  "Name shown on the timeline header")
    config_row(6, "Team Name",      team_name,     "Name of the team")
    config_row(7, "Start Month",    start_month,   "Format: YYYY-MM  (e.g. 2026-04)")
    config_row(8, "Number of Months", num_months,  "How many months the timeline spans")
    config_row(9, "Weeks per Iteration", WEEKS_PER_ITER, "Constant: 2 weeks = 1 iteration")

    # ── Team members ───────────────────────
    section_title(11, "  TEAM MEMBERS")
    styled_cell(ws, 12, 2, "Index", bold=True, font_color=COL_HEADER_FONT,
                fill_color=COL_INFO_LABEL, border=b)
    styled_cell(ws, 12, 3, "Name", bold=True, font_color=COL_HEADER_FONT,
                fill_color=COL_INFO_LABEL, border=b)
    styled_cell(ws, 12, 5, "Role (optional)", bold=True, font_color=COL_HEADER_FONT,
                fill_color=COL_INFO_LABEL, border=b)
    styled_cell(ws, 12, 6, "Email (optional)", bold=True, font_color=COL_HEADER_FONT,
                fill_color=COL_INFO_LABEL, border=b)
    ws.row_dimensions[12].height = 20

    example_roles  = ["Dev", "QA", "Designer", "Scrum Master", "Dev"]
    example_emails = [f"{m.lower()}@company.com" for m in members]

    for i, member in enumerate(members):
        r = 13 + i
        styled_cell(ws, r, 2, i + 1, font_color="1F3864",
                    fill_color="DEEAF1", border=b)
        styled_cell(ws, r, 3, member, font_color="1F3864",
                    fill_color=COL_ITER_ODD, align_h="left", border=b)
        role  = example_roles[i]  if i < len(example_roles)  else ""
        email = example_emails[i] if i < len(example_emails) else ""
        styled_cell(ws, r, 5, role, align_h="left", border=b,
                    fill_color="F2F2F2", font_color="404040")
        styled_cell(ws, r, 6, email, align_h="left", border=b,
                    fill_color="F2F2F2", font_color="404040")
        ws.row_dimensions[r].height = 20

    add_rows = 13 + len(members)
    for extra in range(3):
        r = add_rows + extra
        styled_cell(ws, r, 2, len(members) + extra + 1, font_color="BFBFBF",
                    fill_color="FAFAFA", border=b)
        styled_cell(ws, r, 3, "(add member here)", font_color="BFBFBF",
                    fill_color="FAFAFA", align_h="left", border=b)
        styled_cell(ws, r, 5, "", fill_color="FAFAFA", border=b)
        styled_cell(ws, r, 6, "", fill_color="FAFAFA", border=b)
        ws.row_dimensions[r].height = 20

    last_section_row = add_rows + 3

    # ── Task field legend ──────────────────
    legend_start = last_section_row + 2
    section_title(legend_start, "  TASK FIELD LEGEND")

    fields = [
        ("FEATURE", "Name of the feature / epic this task belongs to",         "e.g.  FEATURE: User Authentication"),
        ("DESCR",   "Short description of what needs to be done",               "e.g.  DESCR: Implement JWT login flow"),
        ("AC",      "Acceptance criterion – when is this task done?",           "e.g.  AC: All unit tests pass, PR merged"),
        ("SP",      "Story Points – effort estimate (Fibonacci: 1,2,3,5,8…)",  "e.g.  SP: 5"),
        ("ITER",    "Iteration number this task is planned for",                "e.g.  ITER: 3"),
    ]

    ws.row_dimensions[legend_start + 1].height = 20
    styled_cell(ws, legend_start + 1, 2, "Field",       bold=True, font_color=COL_HEADER_FONT, fill_color=COL_INFO_LABEL, border=b)
    styled_cell(ws, legend_start + 1, 3, "Meaning",     bold=True, font_color=COL_HEADER_FONT, fill_color=COL_INFO_LABEL, border=b)
    styled_cell(ws, legend_start + 1, 5, "Example",     bold=True, font_color=COL_HEADER_FONT, fill_color=COL_INFO_LABEL, border=b)
    ws.merge_cells(f"E{legend_start+1}:F{legend_start+1}")

    for j, (field, meaning, example) in enumerate(fields):
        r = legend_start + 2 + j
        fill = COL_ITER_ODD if j % 2 == 0 else "FFFFFF"
        styled_cell(ws, r, 2, field,   bold=True,  fill_color=fill, border=b, align_h="left")
        styled_cell(ws, r, 3, meaning, fill_color=fill, border=b, align_h="left", font_size=10)
        ws.merge_cells(f"E{r}:F{r}")
        styled_cell(ws, r, 5, example, fill_color=fill, border=b, align_h="left",
                    font_size=9, font_color="595959")
        ws.row_dimensions[r].height = 20

    # ── Colour legend ──────────────────────
    colour_start = legend_start + 2 + len(fields) + 2
    section_title(colour_start, "  COLOUR LEGEND")
    colour_items = [
        (COL_TEAM_FILL,   COL_TEAM_FONT,   "Team row  – shared tasks visible to everyone"),
        (COL_MEMBER_FILL, COL_MEMBER_FONT, "Individual member row"),
        (COL_TASK_FILL,   "595959",        "Task cell – contains FEATURE / DESCR / AC / SP / ITER"),
        ("D9D9D9",        "404040",        "Iteration header (shading alternates per iteration)"),
    ]
    for k, (bg, fg, desc) in enumerate(colour_items):
        r = colour_start + 1 + k
        styled_cell(ws, r, 2, "■  Sample", bold=True, fill_color=bg,
                    font_color=fg, border=b)
        ws.merge_cells(f"C{r}:F{r}")
        styled_cell(ws, r, 3, desc, fill_color="FAFAFA", align_h="left",
                    border=b, font_size=10)
        ws.row_dimensions[r].height = 20

    # Freeze top rows
    ws.freeze_panes = "B4"

    print("✔  Info & Config sheet built.")


# ─────────────────────────────────────────────
#  SHEET 2 – TIMELINE
# ─────────────────────────────────────────────
def build_timeline_sheet(wb, members, project_name, start_month, num_months, team_name):
    ws = wb.create_sheet("Timeline")
    ws.sheet_view.showGridLines = False

    iterations = calculate_iterations(start_month, num_months)
    num_iters  = len(iterations)

    # Layout: col 1 = fixed width (row labels), then N task-columns per iteration
    LABEL_COL   = 1
    TASK_COLS_PER_ITER = 3        # up to 3 concurrent tasks per iteration per row
    TASK_COL_W  = 24

    # Row layout
    TITLE_ROW    = 1
    MONTH_ROW    = 2              # month band
    ITER_ROW     = 3              # iteration header
    DATE_ROW     = 4              # date range sub-header
    TEAM_ROW     = 5              # team row
    FIRST_MEMBER = 6              # member rows start here

    LABEL_COL_W = 20
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = LABEL_COL_W

    b_header = make_border("medium", "1F3864")
    b_task   = make_border("thin",   COL_TASK_BORDER)
    b_thin   = make_border("thin",   "BFBFBF")
    b_member = make_border("thin",   "6FA8DC")

    # Pre-compute column start positions for each iteration
    iter_col_start = []
    col = LABEL_COL + 1
    for _ in iterations:
        iter_col_start.append(col)
        for tc in range(TASK_COLS_PER_ITER):
            ws.column_dimensions[get_column_letter(col + tc)].width = TASK_COL_W
        col += TASK_COLS_PER_ITER

    total_cols = col - 1

    # ── Row heights ───────────────────────
    ws.row_dimensions[TITLE_ROW].height  = 40
    ws.row_dimensions[MONTH_ROW].height  = 18
    ws.row_dimensions[ITER_ROW].height   = 30
    ws.row_dimensions[DATE_ROW].height   = 16

    TASK_ROW_HEIGHT = 90
    ws.row_dimensions[TEAM_ROW].height   = TASK_ROW_HEIGHT

    for m in range(len(members)):
        ws.row_dimensions[FIRST_MEMBER + m].height = TASK_ROW_HEIGHT

    # ── Title row ─────────────────────────
    ws.merge_cells(start_row=TITLE_ROW, start_column=LABEL_COL,
                   end_row=TITLE_ROW, end_column=total_cols)
    title_cell = ws.cell(row=TITLE_ROW, column=LABEL_COL,
                         value=f"🗓  {project_name}  ·  {team_name}  ·  Timeline")
    title_cell.font      = Font(bold=True, color=COL_HEADER_FONT, size=18)
    title_cell.fill      = make_fill(COL_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Month band ────────────────────────
    from datetime import date, timedelta
    import math
    year, month = map(int, start_month.split("-"))

    # group iterations by month
    month_groups = {}
    for idx, (label, start_str, iter_num) in enumerate(iterations):
        y, m2, d = map(int, start_str.split("-"))
        key = date(y, m2, 1).strftime("%b %Y")
        month_groups.setdefault(key, []).append(idx)

    # Draw month headers (merging consecutive iters in same month)
    from itertools import groupby

    # rebuild as sorted list of (month_label, [iter_indices])
    month_seq = []
    seen = {}
    for idx, (label, start_str, _) in enumerate(iterations):
        y, m2, d = map(int, start_str.split("-"))
        mk = date(y, m2, 1).strftime("%b %Y")
        if mk not in seen:
            seen[mk] = len(month_seq)
            month_seq.append((mk, []))
        month_seq[seen[mk]][1].append(idx)

    for mk, idxs in month_seq:
        c_start = iter_col_start[idxs[0]]
        c_end   = iter_col_start[idxs[-1]] + TASK_COLS_PER_ITER - 1
        ws.merge_cells(start_row=MONTH_ROW, start_column=c_start,
                       end_row=MONTH_ROW, end_column=c_end)
        cell = ws.cell(row=MONTH_ROW, column=c_start, value=mk)
        cell.font      = Font(bold=True, color=COL_HEADER_FONT, size=10)
        cell.fill      = make_fill("2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = b_header

    # Label col for month row
    styled_cell(ws, MONTH_ROW, LABEL_COL, "Month",
                bold=True, font_color=COL_HEADER_FONT,
                fill_color="2F5496", border=b_header)

    # ── Iteration headers ─────────────────
    for idx, (label, start_str, iter_num) in enumerate(iterations):
        cs = iter_col_start[idx]
        ce = cs + TASK_COLS_PER_ITER - 1
        fill_col = COL_ITER_EVEN if idx % 2 == 0 else COL_ITER_ODD

        ws.merge_cells(start_row=ITER_ROW, start_column=cs,
                       end_row=ITER_ROW, end_column=ce)
        iter_label = f"Iter {iter_num}"
        cell = ws.cell(row=ITER_ROW, column=cs, value=iter_label)
        cell.font      = Font(bold=True, color=COL_HEADER_FONT, size=11)
        cell.fill      = make_fill(COL_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = b_header

        # Date sub-header
        y, m2, d = map(int, start_str.split("-"))
        iter_start = date(y, m2, d)
        iter_end   = iter_start + timedelta(days=13)
        date_label = f"{iter_start.strftime('%d %b')} – {iter_end.strftime('%d %b %Y')}"
        ws.merge_cells(start_row=DATE_ROW, start_column=cs,
                       end_row=DATE_ROW, end_column=ce)
        date_cell = ws.cell(row=DATE_ROW, column=cs, value=date_label)
        date_cell.font      = Font(italic=True, color="595959", size=8)
        date_cell.fill      = make_fill(fill_col)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border    = b_thin

    # Label col headers
    styled_cell(ws, ITER_ROW, LABEL_COL, "Iteration",
                bold=True, font_color=COL_HEADER_FONT,
                fill_color=COL_ACCENT, border=b_header)
    styled_cell(ws, DATE_ROW, LABEL_COL, "Sprint dates",
                bold=False, font_color="595959", font_size=8,
                fill_color="F2F2F2", border=b_thin)

    # ── Team row ──────────────────────────
    styled_cell(ws, TEAM_ROW, LABEL_COL, f"🏷  {team_name}\n(Team Tasks)",
                bold=True, font_color=COL_TEAM_FONT,
                fill_color=COL_TEAM_FILL, border=b_member,
                wrap=True, align_h="center")

    TASK_TEMPLATE = "FEATURE: \nDESCR: \nAC: \nSP: \nITER: {iter}"

    for idx, (label, start_str, iter_num) in enumerate(iterations):
        cs        = iter_col_start[idx]
        fill_col  = COL_ITER_EVEN if idx % 2 == 0 else COL_ITER_ODD

        # Merge the 3 sub-columns for team row into groups:
        # First 2 cols = task 1+2, last col = task 3 (or keep separate)
        # Let's keep them separate (each sub-col = one task slot)
        for tc in range(TASK_COLS_PER_ITER):
            c = cs + tc
            if tc == 0:
                task_text = f"FEATURE: \nDESCR: \nAC: \nSP: \nITER: {iter_num}"
            else:
                task_text = ""
            cell = ws.cell(row=TEAM_ROW, column=c, value=task_text)
            cell.font      = Font(color="1F3864", size=9)
            cell.fill      = make_fill(COL_TASK_FILL if tc == 0 else fill_col)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            cell.border    = b_task if tc == 0 else b_thin

    # ── Member rows ───────────────────────
    for mi, member in enumerate(members):
        row = FIRST_MEMBER + mi
        label_fill = COL_MEMBER_FILL if mi % 2 == 0 else "D6E4F0"
        styled_cell(ws, row, LABEL_COL,
                    f"👤  {member}",
                    bold=True, font_color=COL_MEMBER_FONT,
                    fill_color=label_fill, border=b_member,
                    wrap=True, align_h="left")

        for idx, (label, start_str, iter_num) in enumerate(iterations):
            cs       = iter_col_start[idx]
            fill_col = COL_ITER_EVEN if idx % 2 == 0 else COL_ITER_ODD

            for tc in range(TASK_COLS_PER_ITER):
                c = cs + tc
                # Reference team task in first slot via a note; user fills own tasks
                if tc == 0:
                    # Formula reference to team row for same iteration, same col
                    team_cell_addr = f"{get_column_letter(cs)}{TEAM_ROW}"
                    task_text = f"FEATURE: \nDESCR: \nAC: \nSP: \nITER: {iter_num}"
                    note = f"[Team task in {get_column_letter(cs)}{TEAM_ROW}]"
                else:
                    task_text = ""
                    note = ""
                cell = ws.cell(row=row, column=c, value=task_text if tc == 0 else "")
                cell.font      = Font(color="1F3864", size=9)
                cell.fill      = make_fill(COL_TASK_FILL if tc == 0 else fill_col)
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=True)
                cell.border    = b_task if tc == 0 else b_thin

    # ── Column sub-headers (Task 1 / Task 2 / Task 3) ─────────────────
    # Insert a sub-header row just before the team row?
    # We'll add it inside the iteration header as a note in DATE_ROW is already used.
    # Instead, let's add tooltips / comments to iteration first-task cells.

    # ── Freeze panes ──────────────────────
    ws.freeze_panes = ws.cell(row=TEAM_ROW, column=LABEL_COL + 1)

    # ── Print area & zoom ─────────────────
    ws.sheet_view.zoomScale = 85

    print(f"✔  Timeline sheet built  ({num_iters} iterations, {len(members)} members).")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def generate(
    project_name=DEFAULT_PROJECT_NAME,
    team_name=DEFAULT_TEAM_NAME,
    start_month=DEFAULT_START_MONTH,
    num_months=DEFAULT_NUM_MONTHS,
    members=None,
    output_file=OUTPUT_FILE,
):
    if members is None:
        members = DEFAULT_MEMBERS

    wb = openpyxl.Workbook()

    build_info_sheet(wb, members, project_name, start_month, num_months, team_name)
    build_timeline_sheet(wb, members, project_name, start_month, num_months, team_name)

    wb.save(output_file)
    print(f"\n✅  Saved → {output_file}")
    return output_file


if __name__ == "__main__":
    generate()



